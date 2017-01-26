# This code use to get Student and Lecture from Student_Lecture file
from operator import itemgetter

import  openpyxl
import  os
import  datetime
import  uuid
import hashlib

from peewee import *

db = MySQLDatabase('test')
# Link to folder contain excel form
directoryPath = r'C:\Users\HoangHiep\PycharmProjects\Test'

class user(Model) :
    username = CharField(unique=True)
    auth_key = CharField(max_length=32)
    device_hash = CharField(unique=True, null=True)
    password_hash = CharField()
    email = CharField(unique=True)
    profileImg = CharField(null=True)
    status = SmallIntegerField(default=10)
    role = SmallIntegerField(default=10)
    name = CharField(null = True)
    face_id = CharField(max_length=1000, null = True)
    person_id = CharField(null=True)
    created_at = DateTimeField(default=datetime.datetime.now(), null=True)
    updated_at = DateTimeField(null=True)

    class Meta :
        database = db
        
class Student(Model):
    card = CharField(max_length=15, unique= True)
    name = CharField(max_length=255)
    gender = CharField(max_length=1)
    acad = CharField(max_length=10)
    user = ForeignKeyField(user, unique=True)
    created_at = DateTimeField(default = datetime.datetime.now())
    updated_at = DateTimeField(null=True)
    disable = IntegerField(default=0)

    class Meta:
        database = db

class Lecturer(Model):
    card = CharField(max_length=15, unique=True)
    name = CharField(max_length=255)
    acad = CharField(max_length=10)
    email = CharField(max_length=255)
    user = ForeignKeyField(user)
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(null=True)
    disable = IntegerField(default=0)

    class Meta:
        database = db

class venue(Model) :
    location = CharField(max_length=100)
    name = CharField(max_length=100, null=True)
    uuid = CharField(max_length=40, null=True)
    major = SmallIntegerField(null=True)
    minor = SmallIntegerField(null=True)
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(null=True)

    class Meta :
        indexes = (
            ( ('uuid', 'major', 'minor'), True),
        )
        database = db

class Lesson(Model) :
    semester = CharField(max_length=10)
    module_id = CharField(max_length=10)
    subject_area = CharField(max_length=10)
    catalog_number = CharField(max_length=10)
    class_section = CharField(max_length=5)
    component = CharField(max_length=5)
    facility = CharField(max_length=15)
    venue = ForeignKeyField(venue, index=True)
    weekday = CharField(max_length=5)
    start_time = TimeField()
    end_time = TimeField()
    meeting_pattern = CharField(max_length=5, null=True)
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        indexes = (
            (('semester', 'module_id', 'class_section', 'weekday','start_time' ), True),
        )
        database = db

class Lesson_lecturer(Model) :
    lesson = ForeignKeyField(Lesson, index=True)
    lecturer = ForeignKeyField(Lecturer, related_name='Lecturer.id', index=True)
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        indexes = (
            (('lesson', 'lecturer'), True),
        )
        database = db

class Timetable(Model) :
    lesson = ForeignKeyField(Lesson, index=True)
    student = ForeignKeyField(Student, related_name='Student.id', index=True)
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        indexes = (
            (('lesson', 'student'), True),
        )
        database = db

class Lesson_date(Model) :
    lesson = ForeignKeyField(Lesson, index=True)
    ldate = DateField()
    updated_by = IntegerField(default=0, index=True)
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        indexes = (
            (('lesson', 'ldate'), True),
        )
        database = db

class  Attendance(Model) :
    student = ForeignKeyField(Student, index=True)
    lesson_date = ForeignKeyField(Lesson_date, index=True)
    recorded_time = TimeField(default=datetime.datetime.now().time())
    lecturer = ForeignKeyField(Lecturer, index=True)
    status = IntegerField()
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        indexes = (
            (('student', 'lesson_date'), True),
        )
        database = db

class Semester_info(Model) :
    name = CharField(max_length=20)
    start_date = DateField()
    end_date = DateField()
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        database = db

class Semester_date(Model) :
    semester = ForeignKeyField(Semester_info, index=True)
    tdate = DateField(unique=True)
    week_num = IntegerField()
    weekday = IntegerField()
    is_holiday = IntegerField()
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        database = db

class  Beacon_leson(Model):
    lesson = ForeignKeyField(Lesson, unique=True)
    uuid = CharField(default=uuid.uuid4(), unique=True)
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        database = db

class Public_holiday(Model) :
    year = IntegerField()
    name = CharField(max_length=30)
    hdate = DateField()
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(default=datetime.datetime.now())

    class Meta() :
        database = db

class Beacon_attendance_lecturer(Model) :
    lesson_date = ForeignKeyField(Lesson_date, index=True)
    student = ForeignKeyField(Student, index=True)
    lecturer = ForeignKeyField(Lecturer, index=True)
    status = BooleanField()
    created_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        indexes = (
            (('lesson_date', 'student'), True),
        )
        database = db

class Beacon_attendance_student(Model) :
    lesson_date = ForeignKeyField(Lesson_date, index=True)
    student_id_1 = ForeignKeyField(Student, index=True)
    status = BooleanField()
    created_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        indexes = (
            (('lesson_date', 'student_id_1'), True),
        )
        database = db

class Beacon_user(Model) :
    user = ForeignKeyField(user, unique=True)
    major = SmallIntegerField()
    minor = SmallIntegerField()
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        indexes = (
            (('major', 'minor'), True),
        )
        database = db

class Student_leave(Model) :
    student = ForeignKeyField(Student, null=False, index=True)
    start_date = DateField(null=False)
    end_date = DateField(null=False)
    remark = CharField(max_length=100, null=False)
    status = SmallIntegerField(default=0)
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(null=True)

    class Meta :
        indexes = (
            (('student', 'start_date'), True),
        )
        database = db

class  Student_leave_lesson(Model) :
    student = ForeignKeyField(Student, index=True)
    lesson_date = ForeignKeyField(Lesson_date, index=True)
    created_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        indexes = (
            (('student', 'lesson_date'), True),
        )
        database = db

class User_token(Model) :
    user = ForeignKeyField(user, index=True)
    token = CharField(unique=True)
    title = CharField()
    ip_address = CharField()
    expire_date = DateTimeField()
    action = SmallIntegerField()
    created_at = DateTimeField(default=datetime.datetime.now())
    updated_at = DateTimeField(default=datetime.datetime.now())

    class Meta :
        database = db

listLecturer = []
listStudent = []
studentTable = []
lecturerTable = []
listLesson = []
listLessonLecturer = []
listTimetable = []
listAttendance = []
listSemesterdate = []
listLessondate = []
listBeaconLesson = []
listBeaconUser = []
listVenue = []
listUser = []
listStudentLeaveLesson = []

# Connect to database and create new table if table doesn't exist
def initialization() :
    db.connect()
    db.create_tables([user], safe=True)
    db.create_tables([Semester_info], safe=True)
    db.create_tables([Public_holiday], safe=True)
    db.create_tables([Semester_date], safe=True)
    db.create_tables([venue], safe=True)
    db.create_tables([Lesson], safe=True)
    db.create_tables([Lesson_date], safe=True)
    db.create_tables([Student], safe=True)
    db.create_tables([Lecturer], safe=True)
    db.create_tables([Lesson_lecturer], safe=True)
    db.create_tables([Timetable], safe=True)
    db.create_tables([Attendance], safe=True)
    db.create_tables([Beacon_leson], safe=True)
    db.create_tables([User_token], safe=True)
    db.create_tables([Beacon_user], safe=True)
    db.create_tables([Beacon_attendance_lecturer], safe=True)
    db.create_tables([Beacon_attendance_student], safe=True)
    db.create_tables([Student_leave], safe=True)
    db.create_tables([Student_leave_lesson], safe=True)

# Get data from excel form
def getXlsxData() :
    os.chdir(directoryPath)
    folder_list = os.listdir(directoryPath)
    for folders, sub_folders, file in os.walk(directoryPath):
        for name in file :
            if name.endswith(".xlsx") :
                filename = os.path.join(folders, name)
                wb = openpyxl.load_workbook(filename, data_only=True)
                return wb

# Get information of students from Student sheet
def getStudentTable() :
    data = getXlsxData()
    a = data.get_sheet_names()
    students = data.get_sheet_by_name(a[0])
    row_count = students.max_row + 1
    tmp = 1;
    for row in range(3, row_count):
        card = students['B' + str(row)].value
        name = students['C' + str(row)].value
        acad = students['A' + str(row)].value
        gender = 'M'
        moduleID = students['D' + str(row)].value
        offerNbr = students['E' + str(row)].value
        subjectArea = students['F' + str(row)].value
        catalogNbr = students['G' + str(row)].value
        classSection = students['H' + str(row)].value
        component =  students['I' + str(row)].value
        facilityID = students['J' + str(row)].value
        day = students['K' + str(row)].value
        startTime = students['L' + str(row)].value
        endTime = students['M' + str(row)].value
        meetingPattern = students['N' + str(row)].value
        userID = row + tmp
        studentTable.append({
            'card' : card,
            'name' : name,
            'acad' : acad,
            'gender' : gender,
            'moduleID' : moduleID,
            'offerNbr' : offerNbr,
            'subjectArea' : subjectArea,
            'catalogNbr' : catalogNbr,
            'classSection' : classSection,
            'component' : component,
            'facilityID' : facilityID,
            'Day' : day,
            'startTime' : startTime,
            'endTime' : endTime,
            'meetingPattern' : meetingPattern,
            'userID' : userID
        })

# Get information of Lecturers from Lecturer sheet
def getLecturerTable():
    data = getXlsxData()
    a = data.get_sheet_names()
    lecturers = data.get_sheet_by_name(a[1])
    row_count = lecturers.max_row + 1
    tmp = 1
    for row in range(3, row_count):
        userID = tmp + row
        card = lecturers['B' + str(row)].value
        name = lecturers['C' + str(row)].value
        acad = lecturers['A' + str(row)].value
        email = lecturers['D' + str(row)].value
        moduleID = lecturers['E' + str(row)].value
        offerNbr = lecturers['F' + str(row)].value
        subjectArea = lecturers['G' + str(row)].value
        catalogNbr = lecturers['H' + str(row)].value
        classSection = lecturers['I' + str(row)].value
        component = lecturers['J' + str(row)].value
        facilityID = lecturers['K' + str(row)].value
        day = lecturers['L' + str(row)].value
        lecturerTable.append({
            'userID' : userID,
            'card' : card,
            'name' : name,
            'acad' : acad,
            'email' : email,
            'moduleID' : moduleID,
            'offerNbr' : offerNbr,
            'subjectArea' : subjectArea,
            'catalogNbr' : catalogNbr,
            'classSection' : classSection,
            'component' : component,
            'facilityID' : facilityID,
            'day' : day
        })

# Convert information of Student to array list
def getStudent():

    for a in studentTable :
        card = a['card']
        name = a['name']
        acad = a['acad']
        gender = a['gender']
        user_id = a['userID']
        listStudent.append({'card':card, 'name':name, 'gender' : gender, 'acad': acad, 'user':user_id})

# Convert information of Lecturer to array list
def getLecturer() :
    for a in lecturerTable :
        card = a['card']
        name = a['name']
        acad = a['acad']
        email = a['email']
        user_id = a['userID']
        # print(card, name, acad, email, user_id)
        listLecturer.append({'card': card, 'name': name, 'acad': acad, 'email' : email,'user': user_id})

# Validate data of Student list, and delete duplicate Student (base "card" fields)
def validateStudent() :
    _listStudent = sorted(listStudent, key=itemgetter('card'))
    listCard = []

    listStudent.clear()

    for a in _listStudent:
        if (a['card'] not in listCard) :
            listStudent.append(a)
            listCard.append(a['card'])
    # print(len(listStudent))

# update status of Student(leave or not) and add check student in listStudent is new or have appeared at database
def updateStudent() :
    listCard = []
    for a in listStudent:
        if (a['card'] not in listCard):
            listCard.append(a['card'])

    query = Student.update(disable = 1).where(Student.id > 0)
    query.execute()

    query = Student.update(disable=0).where(Student.card.in_(listCard))
    query.execute()

# Add new user into user table
def addUser() :
    if (len(listUser) > 0):
        try:
            with db.atomic():
                user.insert_many(listUser).execute()
        except IntegrityError:
            print('Aaaa')

# Create new user for new Student (user name = card, password = card, email = S + card Number + @connect.np.edu.sg)
def updateStudentUser() :
    for a in listStudent :
        card = a['card']
        t = 'S'
        for i in range(0, len(card) - 1) :
            t = t + str(card[i])
        email = t + '@connect.np.edu.sg'

        hash_object = hashlib.md5(t.encode('utf-8'))
        password = hash_object.hexdigest()
        hash_object = hashlib.md5(email.encode('utf-8'))
        auth_key = hash_object.hexdigest()
        listUser.append(
            {
                'username' : a['card'],
                'auth_key' : auth_key,
                'password_hash' : password,
                'email' : email,
                'name' : a['name'],
            }
        )
    addUser()

    _listUser = user.select()
    for a in _listUser :
        card = a.username
        for b in listStudent :
            if (b['card'] == card) :
                b['user'] = a.id

# add new Student into Student table
def addStudent() :
    count = 0
    _listStudent = sorted(listStudent, key=itemgetter('card'))
    listCurrentStudent = Student
    listCard = []
    for a in listCurrentStudent :
        listCard.append(a.card)
    listStudent.clear()
    # print(len(_listStudent))
    for a in _listStudent :
        if a['card'] not in listCard :
            listStudent.append(a)
    validateStudent()
    updateStudentUser()
    if (len(listStudent) > 0) :
        try :
            with db.atomic():
                Student.insert_many(listStudent).execute()
        except IntegrityError :
            count += 1

# Validate data of Lecturer list, and delete duplicate Lecturer (base "card" fields)
def validateLecturer() :
    _listLecturer = sorted(listLecturer, key=itemgetter('card'))
    listCard = []
    # print(len(_listStudent))
    listLecturer.clear()
    # print(len(_listStudent))
    for a in _listLecturer:
        if (a['card'] not in listCard) :
            listLecturer.append(a)
            listCard.append(a['card'])

# update status of Lecturer(leave or not) and add check Lecturer in listLecturer is new or have appeared at database
def updateLecturer() :
    listCard = []
    for a in listLecturer:
        if (a['card'] not in listCard):
            listCard.append(a['card'])

    query = Lecturer.update(disable = 1).where(Lecturer.id > 0)
    query.execute()
    query = Lecturer.update(disable = 0).where(Lecturer.card.in_(listCard))
    query.execute()

# create new user for new Lecturer
def updateLecturerUser() :
    listUser.clear()
    for a in listLecturer :
        email = a['email']
        t = a['card']
        hash_object = hashlib.md5(t.encode('utf-8'))
        password = hash_object.hexdigest()
        hash_object = hashlib.md5(email.encode('utf-8'))
        auth_key = hash_object.hexdigest()
        listUser.append(
            {
                'username' : email,
                'auth_key' : auth_key,
                'password_hash' : password,
                'email' : email,
                'name' : a['name'],
            }
        )
    addUser()

    _listUser = user.select()
    for a in _listUser :
        email = a.email
        for b in listLecturer :
            if (b['email'] == email) :
                b['user'] = a.id

# add new lecturer to Lecturer table
def addLecturer() :
    count = 0

    _listLecturer = sorted(listLecturer, key=itemgetter('card'))
    listCurrentLecturer = Lecturer
    listCard = []
    for a in listCurrentLecturer:
        listCard.append(a.card)

    listLecturer.clear()
    for a in _listLecturer:
        if a['card'] not in listCard:
            listLecturer.append(a)

    validateLecturer()
    _listLecturer.clear()

    updateLecturerUser()
    if (len(listLecturer) > 0) :
        try :
            with db.atomic():
                Lecturer.insert_many(listLecturer).execute()
        except IntegrityError :
            count += 1
            print('Aaaa')

# get lesson base on Student sheet
def getLesson() :
    for a in studentTable :
        semester = 2
        moduleID = a['moduleID']
        subjectArea = a['subjectArea']
        catalogNumber = a['catalogNbr']
        classSection = a['classSection']
        component = a['component']
        facility = a['facilityID']
        venueID = 1
        weekDay = a['Day']
        startTime = a['startTime']
        endTime = a['endTime']
        meetingPattern = a['meetingPattern']
        listLesson.append({'semester': semester,
                           'module_id': moduleID,
                           'subject_area': subjectArea,
                           'catalog_number': catalogNumber,
                           'class_section': classSection,
                           'component' : component,
                           'facility' : facility,
                           'venue' : venueID,
                           'weekday' : weekDay,
                           'start_time' : startTime,
                           'end_time' : endTime,
                           'meeting_pattern' : meetingPattern
                           })

# Validate and add new lesson into lesson table and reject data if it exist
def addLesson() :
    _listLesson = []
    for a in listLesson :
        if (a not in _listLesson) :
            _listLesson.append(a)
    listLesson.clear()
    for a in _listLesson :
        listLesson.append(a)

    _listLesson.clear()
    _ltt = []
    _listLesson = Lesson.select()

    for a in _listLesson :
        t = str(str(a.semester) + " " + a.module_id +" " +  a.class_section + " "  + a.weekday + " " + str(a.start_time))
        # print(t)
        _ltt.append(t)

    # print(_ltt)
    _listLessont = []
    for a in listLesson :
         t = str(str(a['semester']) + " " + a['module_id'] + " "  + a['class_section'] + " " +  a['weekday'] + " " + str(a['start_time']) + ":00")
         # print(t)
         if (t not in _ltt) :
            _ltt.append(t)
            _listLessont.append(a)

    count = 0

    # print(_listLessont)
    if (len(_listLessont)> 0 ):
        try:
            with db.atomic():
                Lesson.insert_many(_listLessont).execute()
        except IntegrityError:
            count += 1

# Generate lesson lecturer base on Lesson table and Lecturer table
def getLessonLecturer() :
    lesson_list = Lesson.select()
    lecturer_list = Lecturer.select()

    start = datetime.time(0, 30, 0)
    end = datetime.time(22, 59, 0)
    free = []
    d = []
    for i in range (0, len(lesson_list)) :
        d.append(0)

    lecturerEnd = []

    for i in range(0, len(lecturer_list)) :
        free.append(0)
        lecturerEnd.append(datetime.time(0, 0, 0))

    while (start < end) :

        for i in range(0, len(lecturer_list) ):
            if (lecturerEnd[i] <= start) :
                free[i] = 1

        for j in range(0, len(lesson_list)) :
            lesson = lesson_list[j]
            lesson_id = lesson.id
            startLesson = lesson.start_time
            finishLesson = lesson.end_time
            if ((startLesson <= start) and (start <= finishLesson)) :
                if (d[j] == 1) :
                    continue
                for i in range(0, len(lecturer_list)):
                    if (free[i] == 1):
                        free[i] = 0
                        lecturerEnd[i] = finishLesson
                        d[j] = 1
                        listLessonLecturer.append(
                            {
                                'lesson': lesson_id,
                                'lecturer': lecturer_list[i].id
                            }
                        )
                        break
        start = datetime.time(start.hour + 1, start.minute, start.second)

# Add new Lesson Lecturer to lesson_lecturer_table and reject data if it exist
def addLessonLecturer() :
    _lID = Lesson_lecturer.select()
    _list = []
    _listLsLt = []
    for a in _lID :
        _list.append(a.lesson_id)
    for a in listLessonLecturer :
        if (a['lesson'] not in _list) :
            _list.append(a['lesson'])
            _listLsLt.append(a)

    listLessonLecturer.clear()
    for a in _listLsLt :
        listLessonLecturer.append(a)

    count = 0
    if (len(listLessonLecturer) > 0) :
        try:
            with db.atomic():
                Lesson_lecturer.insert_many(listLessonLecturer).execute()
        except IntegrityError:
            count += 1
            print('Aaaa')

# generate Timetable of student base on Lesson table and Student table
def getTimetable() :
    lesson_list = Lesson.select()
    student_list = Student.select()
    free = []
    d = []
    studentEnd = []
    calendar = ['MON', 'TUES', 'WED', 'THUR', 'FRI', 'SAT', 'SUN']
    for wd in range(2, 9) :
        start = datetime.time(0, 30, 0)
        end = datetime.time(22, 59, 0)
        for i in range(0, len(lesson_list)):
            d.append(0)

        for i in range(0, len(student_list)):
            free.append(0)
            studentEnd.append(datetime.time(0, 0, 0))

        while (start < end):

            for i in range(0, len(student_list)):
                if (studentEnd[i] <= start):
                    free[i] = 1

            for j in range(0, len(lesson_list)):
                lesson = lesson_list[j]
                lesson_id = lesson.id
                startLesson = lesson.start_time
                finishLesson = lesson.end_time
                weekday = lesson.weekday
                # print(weekday, wd)
                if (weekday == calendar[wd - 2]) :
                    if ((startLesson <= start) and (start <= finishLesson)):
                        if (d[j] == 1):
                            continue
                        for i in range(0, len(student_list)):
                            if (free[i] == 1):
                                free[i] = 0
                                studentEnd[i] = finishLesson
                                d[j] = 1
                                listTimetable.append(
                                    {
                                        'lesson': lesson_id,
                                        'student': student_list[i].id
                                    }
                                )
                                break
            start = datetime.time(start.hour + 1, start.minute, start.second)

# add new Timetable to timetable_table and reject data if it exist
def addTimetable() :
    count = 0
    _l = []
    _listTimeTable = []
    _t = Timetable.select()
    for a in _t :
        _l.append(a.lesson_id * 10000 + a.student_id)

    for a in listTimetable :
        t = a['lesson'] * 10000 + a['student']
        if (t not in _l) :
            _listTimeTable.append(a)

    if (len(_listTimeTable) > 0) :
        try:
            with db.atomic():
                Timetable.insert_many(_listTimeTable).execute()
        except IntegrityError:
            count += 1
            print('Aaaa')

# Generate Attendance base on lesson_date, lesson, student table
def getAttendance() :
    lsdate = Lesson_date.select()
    timetb = Timetable.select()
    lslec = Lesson_lecturer.select()
    for a in lsdate :
        lsID = a.lesson_id
        lcID = 0
        for b in lslec :
            if (b.lesson_id == lsID):
                lcID = b.lecturer_id

        for b in timetb :
            if (lsID == b.lesson_id):
                listAttendance.append(
                    {
                        'student' : b.student_id,
                        'lesson_date' : a.id,
                        'recorded_time' : datetime.datetime.now().time(),
                        'lecturer' : lcID,
                        'status' : lsID
                    }
                )

# Insert attendance list to database and reject data if it exist
def addAttendance() :
    count = 0
    _listAttendance = []
    _l = []
    _t = Attendance.select()
    for a in _t :
        t = str(a.student_id) + " " + str(a.lesson_date_id)
        _l.append(t)

    for a in listAttendance :
        t = str(a['student']) + " " + str(a['lesson_date'])
        if (t not in _l) :
            _listAttendance.append(a)

    if (len(_listAttendance) > 0) :
        try:
            with db.atomic():
                Attendance.insert_many(_listAttendance).execute()
        except IntegrityError:
            count += 1
            print('Aaaa')

# Generate [semester dates] after [public holiday] and [semster info] tables are setup
def getSemesterdate():
    public_holiday = Public_holiday.select()
    semester_date = Semester_date.select()
    semester_info = Semester_info.select()
    list_holiday = [];
    for a in public_holiday:
        list_holiday.append(a.hdate)

    list_semester_id = []
    for a in semester_date:
        list_semester_id.append(a.semester_id)
    list_semester_id = list(set(list_semester_id))

    for a in semester_info:
        semester_id = a.id
        if (semester_id not in list_semester_id):
            start_date = a.start_date
            end_date = a.end_date
            week_num = 0;
            while start_date <= end_date:
                tdate = start_date;
                weekday = tdate.isoweekday()
                week_num += (weekday == 1)
                start_date += datetime.timedelta(days=1)
                if tdate in list_holiday:
                    is_holiday = 1
                else:
                    is_holiday = 0
                listSemesterdate.append(
                    {
                        'semester' : semester_id,
                        'tdate' : tdate,
                        'week_num' : week_num,
                        'weekday' : weekday,
                        'is_holiday' : is_holiday
                    }
                )

#  Insert list of Semesterdates  to database and reject data if it exist
def addSemesterdate() :
    count = 0
    if (len(listSemesterdate) > 0) :
        try:
            with db.atomic():
                Semester_date.insert_many(listSemesterdate).execute()
        except IntegrityError:
            count += 1
            print('Aaaa')

# Generate [lesson dates] after [semester date] and [lesson] tables are set up
def getLessondate() :
    list_lesson = Lesson.select()
    list_semester_date = Semester_date.select()
    list_lesson_date = Lesson_date.select()
    calendar = ['MON', 'TUES', 'WED', 'THUR', 'FRI', 'SAT', 'SUN']

    list_ldate = []
    for a in list_lesson_date:
        list_ldate.append(a.ldate)
    list_ldate = list(set(list_ldate))

    for a in list_semester_date:
        weekday = a.weekday
        # print(a.weekday)
        ldate = a.tdate
        if (a.is_holiday == 0) :
            if (ldate not in list_ldate):
                for b in list_lesson:
                    weekday_ = calendar[weekday - 1]
                    if (weekday_ == b.weekday) and (weekday < 6):
                        lesson_id = b.id
                        ldate = a.tdate
                        updated_by = 2;
                        listLessondate.append(
                            {
                                'lesson' : lesson_id,
                                'ldate' : ldate,
                                'updated_by' : updated_by
                            }
                        )

# Insert list of Lessondate  to database and reject data if it exist
def addLessondate() :
    count = 0
    # print(listLessondate)
    if (len(listLessondate) > 0) :
        try:
            with db.atomic():
                Lesson_date.insert_many(listLessondate).execute()
        except IntegrityError:
            count += 1
            print('Aaaa')

# Generate uuid for each beacon lesson
def getBeaconLesson():
    BcL = Beacon_leson.select()
    _num = 10000000;
    _l = []
    for a in BcL :
        if (a.lesson_id not in _l) :
            _l.append(a.lesson_id)

    listLs = Lesson.select()
    for a in listLs :
        if (a.id not in _l) :
            _l.append(a)
            _num += 1
            listBeaconLesson.append(
                {
                    'lesson' : a.id,
                    'uuid' : uuid.uuid4()
                }
            )

# Insert Beaconlesson list to database and reject data if it exist
def addBeaconLesson():
    if (len(listBeaconLesson) > 0):
        try:
            with db.atomic():
                Beacon_leson.insert_many(listBeaconLesson).execute()
        except IntegrityError:
            print('Aaaa')

# Generate uuid, major, minor for Beacon user
def getBeaconUser() :
    BcU = Beacon_user.select()
    l = []
    for a in BcU :
        if (a.user_id not in l) :
            l.append(a.user_id)

    listUser = user.select()
    _minor = len(l) + 1;
    _major = 99 - len(l) - 1;
    for a in listUser :
        if (a.id not  in l ) :
            l.append(a)
            _minor += 1
            _major -= 1
            listBeaconUser.append(
                {
                    'user' : a.id,
                    'major' : _major,
                    'minor' : _minor
                }
            )

# Insert BeaconUser list to database and reject data if it exist
def addBeaconUser() :
    count = 0
    if (len(listBeaconUser) > 0) :
        try:
            with db.atomic():
                Beacon_user.insert_many(listBeaconUser).execute()
        except IntegrityError:
            count += 1
            print('Aaaa')

# Generate venue data(location, name)
def getVenue() :
    _l = []
    for a in studentTable :
        _l.append(a['facilityID'])

    _listVenue = list(set(_l))
    for a in _listVenue :
        listVenue.append(
            {
                'location' : a,
                'name' : a
            }
        )

# Add venue list to database and reject data if it exist
def addVenue() :
    if (len(listVenue) > 0):
        try:
            with db.atomic():
                venue.insert_many(listVenue).execute()
        except IntegrityError:
            print('Aaaa')

# Generate [StudentLeaveLesson] base on [Student_leave] table and [Lesson] table
# tetsfdsf
def getStudentLeaveLesson() :
    listStudentLeave = Student_leave.select()
    t = Timetable.select()
    listLsDate = Lesson_date.select()
    listLs = []
    for a in listStudentLeave :
        studentID = a.student_id
        listLs.clear()
        start_date = a.start_date
        end_date = a.end_date
        for b in t :
            if (b.student_id == studentID) :
                listLs.append(b.lesson_id)
        for a in listLs :
            lesson_id  = a
            for b in listLsDate :
                date = b.ldate
                if ((start_date <= date) and (date <= end_date)) :
                    if (b.lesson_id == lesson_id) :
                        listStudentLeaveLesson.append(
                            {
                                'student' : studentID,
                                'lesson_date' : b.id
                            }
                        )

# Import StudentLeaveLesson list to database and reject data if it exist
def addStudentLeaveLesson() :
    _lsll = Student_leave_lesson.select()
    _lls = []
    for a in _lsll :
        t = str(a.student_id) + " " + str(a.lesson_date_id)
        if (t not in _lls) :
            _lls.append(t)
    _listStudentLeaveLesson = []
    for a in listStudentLeaveLesson :
        t = str(a['student']) + " " + str(a['lesson_date'])
        if (t not in _lls) :
            _lls.append((a))
            _listStudentLeaveLesson.append(a)

    listStudentLeaveLesson.clear()
    for a in _listStudentLeaveLesson :
        listStudentLeaveLesson.append(a)

    if (len(listStudentLeaveLesson) > 0) :
        try :
            with db.atomic():
                Student_leave_lesson.insert_many(listStudentLeaveLesson).execute()
        except IntegrityError :
            print('AAA')

def importStudent() :
    getStudent()
    updateStudent()
    addStudent()

def importLecturer():
    getLecturer()
    updateLecturer()
    addLecturer()

def importLesson() :
    getLesson()
    addLesson()

def importVenue() :
    getVenue()
    addVenue()

def importLessonLecturer() :
    getLessonLecturer()
    addLessonLecturer()

def importTimetable() :
    getTimetable()
    addTimetable()

def importAttendance() :
    getAttendance()
    addAttendance()

def generateSemesterdate() :
    getSemesterdate()
    addSemesterdate()

def generateLessondate() :
    getLessondate()
    addLessondate()

def generateBeaconLesson() :
    getBeaconLesson()
    addBeaconLesson()

def generateStudentLeaveLesson():
    getStudentLeaveLesson()
    addStudentLeaveLesson()

def generateBeaconUser() :
    getBeaconUser()
    addBeaconUser()


if __name__ == '__main__' :
    print('Start')
    initialization()
    getStudentTable()
    getLecturerTable()
    importVenue()
    importStudent()
    importLecturer()
    generateSemesterdate()
    importLesson()
    generateLessondate()
    generateBeaconUser()
    generateBeaconLesson()
    importLessonLecturer()
    importTimetable()
    importAttendance()
    generateStudentLeaveLesson()
    print('Finish')